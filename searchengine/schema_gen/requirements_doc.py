"""要件定義書テンプレート生成モジュール（Markdown / Excel）。

Markdown: 完全な要件定義書テンプレート（ヒアリングシート・機能一覧・非機能要件含む）
Excel:    openpyxl で複数シート構成の要件定義書を生成
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from .analyzer import ColumnInfo
from .er import detect_fk_candidates, detect_normalization_issues


def _today() -> str:
    return date.today().isoformat()


# ---------------------------------------------------------------------------
# Markdown 出力
# ---------------------------------------------------------------------------

def render_requirements_md(
    filename: str,
    table_name: str,
    columns: list[ColumnInfo],
    row_count: int,
) -> str:
    pk_cols = [c for c in columns if c.pk_candidate]
    nullable_cols = [c for c in columns if c.null_pct > 0]
    fk_relations = detect_fk_candidates(columns, table_name)
    norm_hints = detect_normalization_issues(columns)
    today = _today()

    lines = [
        f"# 要件定義書 — {table_name}",
        "",
        f"> **自動生成**: `{filename}` を解析して生成（{today}）  ",
        "> ⚠️ このドキュメントはたたき台です。内容を確認・修正してから使用してください。",
        "",
        "---",
        "",
        "## 1. プロジェクト概要",
        "",
        "| 項目 | 内容 |",
        "|---|---|",
        f"| システム名 | （記入） |",
        f"| 対象テーブル | `{table_name}` |",
        f"| 作成日 | {today} |",
        f"| 作成者 | （記入） |",
        f"| バージョン | 0.1 |",
        "",
        "### 背景・目的",
        "",
        "> （クライアントへのヒアリング内容を記入）",
        "",
        "---",
        "",
        "## 2. データ概要",
        "",
        f"- **ソースファイル**: `{filename}`",
        f"- **行数**: {row_count:,}",
        f"- **カラム数**: {len(columns)}",
        "",
        "### カラム定義",
        "",
        "| # | カラム名 | 型 | NULL | 主キー | 説明 |",
        "|---|---|---|---|---|---|",
    ]

    for i, c in enumerate(columns, 1):
        pk = "✅" if c.pk_candidate else ""
        null = "○" if c.null_pct > 0 else "✕"
        lines.append(f"| {i} | `{c.name}` | {c.inferred_type} | {null} | {pk} | （記入） |")

    if fk_relations:
        lines += [
            "",
            "### リレーション（自動推定）",
            "",
            "| カラム | 参照先テーブル | 備考 |",
            "|---|---|---|",
        ]
        for r in fk_relations:
            lines.append(f"| `{r.from_col}` | `{r.ref_table}.{r.ref_col}` | 要確認 |")

    lines += [
        "",
        "---",
        "",
        "## 3. 機能要件",
        "",
        "### 3-1. CRUD 操作",
        "",
        "| 操作 | 必須 | 条件・補足 |",
        "|---|---|---|",
        "| 新規登録 | ☐ | |",
        "| 一覧表示 | ☐ | ページング: ☐ |",
        "| 詳細表示 | ☐ | |",
        "| 編集・更新 | ☐ | |",
        "| 削除 | ☐ | 論理削除: ☐ / 物理削除: ☐ |",
        "| 一括インポート | ☐ | CSV / Excel |",
        "| エクスポート | ☐ | CSV / Excel / PDF |",
        "",
        "### 3-2. 検索・絞り込み",
        "",
        "| カラム | 検索方式 | 必須 |",
        "|---|---|---|",
    ]

    for c in columns[:8]:  # 代表的なカラムのみ列挙
        lines.append(f"| `{c.name}` | 完全一致 / 部分一致 / 範囲 | ☐ |")

    lines += [
        "",
        "### 3-3. バリデーション",
        "",
        "| カラム | ルール | エラーメッセージ |",
        "|---|---|---|",
    ]

    for c in columns:
        null_rule = "任意" if c.null_pct > 0 else "必須"
        type_hint = "数値のみ" if "INT" in c.inferred_type or "NUMERIC" in c.inferred_type else "文字列"
        lines.append(f"| `{c.name}` | {null_rule} / {type_hint} | （記入） |")

    lines += [
        "",
        "---",
        "",
        "## 4. 非機能要件",
        "",
        "| 項目 | 要件 |",
        "|---|---|",
        "| パフォーマンス | 一覧表示: ___秒以内 / 検索: ___秒以内 |",
        "| 同時アクセス | ___ユーザー |",
        "| データ保持期間 | ___年 |",
        "| バックアップ | 頻度: ___ / 保持: ___ |",
        "| セキュリティ | 認証: ☐ / 権限管理: ☐ / 暗号化: ☐ |",
        "| ブラウザ対応 | Chrome / Firefox / Safari / Edge |",
        "| モバイル対応 | 不要 / レスポンシブ / ネイティブアプリ |",
        "",
        "---",
        "",
        "## 5. ヒアリング項目",
        "",
    ]

    if nullable_cols:
        lines.append("### 任意項目の扱い")
        lines.append("")
        for c in nullable_cols:
            lines.append(f"- [ ] `{c.name}`（NULL率 {c.null_pct}%）— 空になるケースは？必須化できるか？")
        lines.append("")

    if norm_hints:
        lines.append("### 設計上の確認事項")
        lines.append("")
        for h in norm_hints:
            lines.append(f"- [ ] {h.issue}")
            lines.append(f"  → {h.suggestion}")
        lines.append("")

    lines += [
        "### その他",
        "",
        "- [ ] 既存システムとの連携はあるか？",
        "- [ ] データ移行（マイグレーション）は必要か？",
        "- [ ] 承認フロー・ワークフローは必要か？",
        "- [ ] 通知機能（メール・Slack等）は必要か？",
        "- [ ] 多言語対応は必要か？",
        "",
        "---",
        "",
        "## 6. 未決事項",
        "",
        "| # | 内容 | 担当 | 期限 |",
        "|---|---|---|---|",
        "| 1 | | | |",
        "| 2 | | | |",
        "",
        "---",
        "",
        f"*このドキュメントは `schema-gen` によって自動生成されました（{today}）*",
    ]

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Excel 出力
# ---------------------------------------------------------------------------

def render_requirements_xlsx(
    filename: str,
    table_name: str,
    columns: list[ColumnInfo],
    row_count: int,
    out_path: Path,
) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl が必要です: pip install openpyxl")

    wb = openpyxl.Workbook()
    today = _today()

    # ---- シート1: 概要 ----
    ws1 = wb.active
    ws1.title = "概要"
    _header_row(ws1, 1, "要件定義書", table_name)
    ws1.append([])
    ws1.append(["項目", "内容"])
    for row in [
        ["システム名", "（記入）"],
        ["対象テーブル", table_name],
        ["ソースファイル", filename],
        ["行数", row_count],
        ["カラム数", len(columns)],
        ["作成日", today],
        ["作成者", "（記入）"],
        ["バージョン", "0.1"],
    ]:
        ws1.append(row)
    _style_header(ws1, 3, 2)
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 40

    # ---- シート2: カラム定義 ----
    ws2 = wb.create_sheet("カラム定義")
    headers = ["#", "カラム名", "型", "NULL率", "一意率", "PK候補", "サンプル値", "説明・備考"]
    ws2.append(headers)
    _style_header(ws2, 1, len(headers))
    for i, c in enumerate(columns, 1):
        ws2.append([
            i,
            c.name,
            c.inferred_type,
            f"{c.null_pct}%",
            f"{c.unique_pct}%",
            "✅" if c.pk_candidate else "",
            ", ".join(c.sample_values[:3]),
            "",
        ])
    for col, width in zip("ABCDEFGH", [4, 28, 20, 10, 10, 8, 30, 30]):
        ws2.column_dimensions[col].width = width

    # ---- シート3: 機能要件 ----
    ws3 = wb.create_sheet("機能要件")
    ws3.append(["操作", "必須", "条件・補足"])
    _style_header(ws3, 1, 3)
    for op in ["新規登録", "一覧表示", "詳細表示", "編集・更新", "削除", "一括インポート", "エクスポート"]:
        ws3.append([op, "☐", ""])
    ws3.append([])
    ws3.append(["カラム", "検索方式", "必須"])
    _style_header(ws3, ws3.max_row, 3)
    for c in columns[:8]:
        ws3.append([c.name, "完全一致 / 部分一致 / 範囲", "☐"])
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["C"].width = 10

    # ---- シート4: ヒアリング ----
    ws4 = wb.create_sheet("ヒアリング")
    ws4.append(["確認項目", "回答", "担当", "期限"])
    _style_header(ws4, 1, 4)
    fk_relations = detect_fk_candidates(columns, table_name)
    norm_hints = detect_normalization_issues(columns)
    nullable_cols = [c for c in columns if c.null_pct > 0]
    for c in nullable_cols:
        ws4.append([f"`{c.name}` が空になるケースは？（NULL率 {c.null_pct}%）", "", "", ""])
    for h in norm_hints:
        ws4.append([h.suggestion, "", "", ""])
    for item in ["既存システムとの連携はあるか？", "データ移行は必要か？", "承認フローは必要か？"]:
        ws4.append([item, "", "", ""])
    for col, width in zip("ABCD", [50, 30, 15, 15]):
        ws4.column_dimensions[col].width = width

    wb.save(out_path)


def _header_row(ws, row: int, title: str, subtitle: str = "") -> None:
    ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=14)
    if subtitle:
        ws.cell(row=row, column=2, value=subtitle)


def _style_header(ws, row: int, ncols: int) -> None:
    try:
        from openpyxl.styles import Font, PatternFill
        fill = PatternFill("solid", fgColor="4472C4")
        font = Font(bold=True, color="FFFFFF")
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.font = font
    except Exception:
        pass
