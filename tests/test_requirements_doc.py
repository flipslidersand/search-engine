"""requirements_doc.py のテスト"""
import pytest
from pathlib import Path
from searchengine.schema_gen.analyzer import analyze
from searchengine.schema_gen.requirements_doc import render_requirements_md, render_requirements_xlsx


def _make_rows(n=5):
    return [
        {"order_id": str(i), "customer_id": str(i % 3 + 1), "total": str(i * 100), "note": None}
        for i in range(1, n + 1)
    ]


@pytest.fixture
def cols():
    return analyze(_make_rows())


class TestRenderRequirementsMd:
    def test_returns_string(self, cols):
        md = render_requirements_md("test.csv", "orders", cols, 5)
        assert isinstance(md, str)

    def test_has_heading(self, cols):
        md = render_requirements_md("test.csv", "orders", cols, 5)
        assert "# 要件定義書" in md
        assert "orders" in md

    def test_has_column_table(self, cols):
        md = render_requirements_md("test.csv", "orders", cols, 5)
        assert "カラム定義" in md
        assert "order_id" in md
        assert "customer_id" in md

    def test_has_sections(self, cols):
        md = render_requirements_md("test.csv", "orders", cols, 5)
        for section in ["機能要件", "非機能要件", "ヒアリング", "未決事項"]:
            assert section in md

    def test_nullable_column_in_hearing(self, cols):
        md = render_requirements_md("test.csv", "orders", cols, 5)
        # note カラムは NULL あり → ヒアリングセクションに出る
        assert "note" in md

    def test_fk_relation_shown(self, cols):
        md = render_requirements_md("test.csv", "orders", cols, 5)
        # customer_id は FK 候補 → リレーション節に表示される
        assert "customers" in md

    def test_row_count_shown(self, cols):
        md = render_requirements_md("test.csv", "orders", cols, 1234)
        assert "1,234" in md

    def test_filename_shown(self, cols):
        md = render_requirements_md("myfile.csv", "tbl", cols, 5)
        assert "myfile.csv" in md


class TestRenderRequirementsXlsx:
    def test_creates_file(self, cols, tmp_path):
        pytest.importorskip("openpyxl")
        out = tmp_path / "req.xlsx"
        render_requirements_xlsx("test.csv", "orders", cols, 5, out)
        assert out.exists()
        assert out.stat().st_size > 0

    def test_has_four_sheets(self, cols, tmp_path):
        pytest.importorskip("openpyxl")
        import openpyxl
        out = tmp_path / "req.xlsx"
        render_requirements_xlsx("test.csv", "orders", cols, 5, out)
        wb = openpyxl.load_workbook(out)
        assert set(wb.sheetnames) == {"概要", "カラム定義", "機能要件", "ヒアリング"}

    def test_column_names_in_sheet(self, cols, tmp_path):
        pytest.importorskip("openpyxl")
        import openpyxl
        out = tmp_path / "req.xlsx"
        render_requirements_xlsx("test.csv", "orders", cols, 5, out)
        wb = openpyxl.load_workbook(out)
        ws = wb["カラム定義"]
        values = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        assert "order_id" in values
        assert "customer_id" in values

    def test_raises_import_error_without_openpyxl(self, cols, tmp_path, monkeypatch):
        import builtins
        real_import = builtins.__import__

        def mock_import(name, *args, **kwargs):
            if name == "openpyxl":
                raise ImportError("mocked")
            return real_import(name, *args, **kwargs)

        monkeypatch.setattr(builtins, "__import__", mock_import)
        with pytest.raises(ImportError):
            render_requirements_xlsx("test.csv", "orders", cols, 5, tmp_path / "x.xlsx")
